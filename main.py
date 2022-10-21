import openpyxl
from bs4 import BeautifulSoup  # 网页解析，获取数据
import re  # 正则表达式，进行文字匹配`
import urllib.request, urllib.error  # 制定URL，获取网页数据
import xlwt  # 进行excel操作
import xlsxwriter
import xlrd3
import openpyxl
import lxml
import re
import html5lib
from scrapy import Selector
# import sqlite3  # 进行SQLite数据库操作
import requests

def main():
    ligne=0
    for i in range(0,10):
        baseurl = "https://www.huarenjiewang.com/category-23-9-"+str(i+1)+".html"  # 要爬取的网页链接
        # 1.爬取网页

        datatitlejob = getData(baseurl,'title')
        datatitlejob=datatitlejob [0:30]
        datalinkjob = getData(baseurl,'link')
        datadate = getData(baseurl, 'date')
        savepath = "huarenjie.xlsl"  # 当前目录新建XLS，存储进去
        # 3.保存数据
        saveData(datatitlejob,5, ligne)  # 存储
        saveData(datadate, 9, ligne)
        #saveData(datalinkjob,10, ligne)
        if len(datatitlejob)==len(datalinkjob):
            ligne=ligne + len(datatitlejob)
        else:
            print("error")
        for j in range(len(datalinkjob)):
            baseurl=datalinkjob[j]
            datainfo=getData(baseurl,'info')
            saveData(datainfo, 1, j+i*30)

def readexcel(data,clon):
    data=[]
    workbook = xlrd3.open_workbook('huarenjie.xlsx')
    SheetNameList = workbook.sheet_names()
    print(SheetNameList[0])
    worksheet = workbook.sheet_by_name(SheetNameList[0])
    num_rows = worksheet.nrows
    num_cells = worksheet.ncols
    print('num_rows, num_cells', num_rows, num_cells)
    curr_row = 1

    while curr_row < num_rows:
        row = worksheet.row(curr_row)
        # print row, len(row), row[0], row[1]
        print('Row: ', curr_row)
        print(row, len(row), row[0])
        curr_cell = 0
        while curr_cell < num_cells:
            # Cell Types: 0=Empty, 1=Text, 2=Number, 3=Date, 4=Boolean, 5=Error, 6=Blank
            cell_type = worksheet.cell_type(curr_row, curr_cell)
            cell_value = worksheet.cell_value(curr_row, curr_cell)
            print(' ', cell_type, ':', cell_value)
            curr_cell += 1
        curr_row += 1

    return data
# 爬取网页

def getData(baseurl,param):
    baseurl=baseurl.strip('"')
    print(baseurl)
    datalist = []  # 用来存储爬取的网页信息
    html = askURL(baseurl)


    soup= BeautifulSoup(html, "html.parser")
    if param =='title':
        for item in soup.find_all(class_="ltitle"):
            #print(item.text)
            datalist.append(item.text)
    elif param == 'date':
        k=0
        for item in soup.find_all(class_="ltime"):
            if k%3==2:
                datalist.append(item.text)
            k+=1

    elif param =='link':
        for link in soup.find_all(class_="ltitle"):
            # print(str(link.a).strip('<a href= ')[1:49])
            datalist.append(str(link.a).strip('<a href= ')[1:49])
    elif param =='info':
        #print(soup)
        typejob = ""
        categjob = ""
        phone = ""
        c3 = soup.find(class_="contact")
        c3 = c3.find_all("span")
        sujet=[]
        info=[]
        index=[]
        data=[]
        c4 = soup.find(class_="contact")
        c4 = c4.find_all(style="")
        for i in range(len(c3)):
            info.append(str(c3[i]).strip('<span> ：</ '))
            index.append(str(c4).find(info[i]))
            data.append(str(c4)[index[i]+10:index[i]+15].strip('n> < fo'))
            sujet.append((c3[i]).text.strip())
        print(sujet)

        if soup.find(class_="contact") is not None:
            c1 = soup.find(class_="contact")
            c1 = c1.find_all("li", style="")
            indtype=findindex('性质：',sujet)
            if indtype != -1:
                type = c1[indtype].find_all(class_="mayi")
                for i in range(len(type)):
                    typejob += type[i].text + " "


            indcateg = findindex('工种：', sujet)
            if indcateg != -1:
                if len(c1)>3 and c1[indcateg].find_all(class_="mayi")!="":
                    categ = c1[indcateg].find_all(class_="mayi")
                    for i in range(len(categ)):
                        categjob += categ[i].text + " "

        if soup.find(id="show-phone") is not None:
            c2 = soup.find(id="show-phone")
            print(c2)
            phone=re.findall('\d+', str(c2))[0]
            i=1
            for i in range(len(re.findall('\d+', str(c2)))):
                if len(phone)<10 :
                    phone += re.findall('\d+', str(c2))[i]
            print(phone[0:10])

        contenu=soup.find(class_="view_bd").text.strip()
        indzone=findindex('区域：',sujet)
        if indzone != -1:
            zone=data[indzone].strip('n\ </>')
        else:
            zone=""
        indoffre= findindex('供求：', sujet)
        if indoffre != -1:
            offre=data[indoffre].strip('n\ </>')
        else:
            offre=""


        domaines = soup.find(class_="location")
        for item in domaines.find_all("a"):
            domaine=item.text
        print(domaine)

        datalist.append("infos")
        datalist.append(offre)
        datalist.append(domaine)
        if categjob!="":
            datalist.append(categjob)
            datalist.append(typejob)
        else:
            datalist.append(typejob)
            datalist.append(categjob)

        datalist.append(contenu)
        datalist.append(zone)
        datalist.append(phone)

        print(datalist)
    else:
        print("error")
    return datalist



# 得到指定一个URL的网页内容
def askURL(url):
    head = {  # 模拟浏览器头部信息，向服务器发送消息
        "User-Agent": "Mozilla / 5.0(Windows NT 10.0; Win64; x64) AppleWebKit / 537.36(KHTML, like Gecko) Chrome / 80.0.3987.122  Safari / 537.36"
    }

    request = urllib.request.Request(url, headers=head)
    html = ""
    try:
        response = urllib.request.urlopen(request)
        html = response.read().decode("utf-8", "ignore")
    except urllib.error.URLError as e:
        if hasattr(e, "code"):
            print(e.code)
        if hasattr(e, "reason"):
            print(e.reason)
    return html


# 保存数据到表格
def saveData(datalist, colon, ligne):
    print("save.......")
    workbook = openpyxl.load_workbook('huarenjie.xlsx', read_only=False)
    worksheet = workbook.active
    if datalist[0]=="infos":
        k=0
        #print(datalist[0], len(datalist))
        for i in range(len(datalist)-1):
            if i==4:
                k+=1
            worksheet.cell(ligne+2, colon+i+k).value = datalist[i+1]
    else:
        for i in range(len(datalist)):
            worksheet.cell(i+ligne+2, colon).value = datalist[i]
    workbook.save("huarenjie.xlsx")

# 保存数据到数据库

def findindex(mot, list):
    if mot in list:
        return list.index(mot)
    else:
        return -1

if __name__ == "__main__":  # 当程序执行时
    # 调用函数
    main()
    # init_db("movietest.db")
    print("爬取完毕！")
